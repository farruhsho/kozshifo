"""Pure builders that turn a (title, header, rows) report table into downloadable
XLSX and PDF bytes — the binary counterparts of the CSV export in
`features/reports.py`.

Both builders are presentation-only: the caller passes the SAME header + rows it
already feeds the CSV writer, so column content/order stays identical across all
three formats.

Cyrillic: the PDF reuses the bundled DejaVu font that `core.print_forms` already
registers (reportlab's built-in Helvetica has no Cyrillic glyphs and prints
«чёрные квадратики»). XLSX delegates glyph rendering to the spreadsheet app, so
no font handling is needed there.
"""
from __future__ import annotations

import io
from decimal import Decimal
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle


def _xlsx_cell(value: object) -> object:
    """Decimal/float → float (so Excel treats it as a number); everything else
    passes through unchanged (None becomes an empty cell)."""
    if isinstance(value, Decimal):
        return float(value)
    return value


def build_xlsx(title: str, header: Sequence[str], rows: Sequence[Sequence]) -> bytes:
    """One worksheet: optional bold title row, a bold frozen header row, then the
    data rows. Numbers are written as numbers; column widths fit the content."""
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]  # Excel caps sheet names at 31 chars

    col_count = len(header)
    title_offset = 0
    if title:
        ws.append([title])
        cell = ws.cell(row=1, column=1)
        cell.font = Font(bold=True, size=13)
        if col_count > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_offset = 1

    header_row_idx = title_offset + 1
    ws.append(list(header))
    for col in range(1, col_count + 1):
        c = ws.cell(row=header_row_idx, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append([_xlsx_cell(v) for v in row])

    # Freeze everything above the first data row (title + header stay on screen).
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    # Column widths: widest of header / cell text, clamped to something sane.
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        width = len(str(header[col - 1]))
        for row in rows:
            if col - 1 < len(row):
                width = max(width, len(str(row[col - 1])))
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(title: str, header: Sequence[str], rows: Sequence[Sequence]) -> bytes:
    """A4-landscape table: title paragraph + a [header]+rows table with a bold
    grey header that repeats on page breaks. Uses the bundled DejaVu font so
    Cyrillic renders (no tofu)."""
    # Reuse print_forms' DejaVu registration — guarantees Cyrillic glyphs.
    from app.core import print_forms
    print_forms._register_fonts()
    font, font_bold = print_forms.FONT, print_forms.FONT_BOLD

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=title or "Report",
    )

    title_style = ParagraphStyle("ReportTitle", fontName=font_bold, fontSize=14, spaceAfter=8)
    elements = [Paragraph(str(title or ""), title_style)]

    data = [[str(h) for h in header]] + [[str(c) for c in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTNAME", (0, 0), (-1, 0), font_bold),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()
